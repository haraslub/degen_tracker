from openpyxl import Workbook, load_workbook, workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule

from config import _dict_columns_to_hide, dict_columns_width, adjust_dict

import os
import pandas as pd
import string


def xlsx_update(filepath, sheet_name, list_columns, dt):
    wb = load_workbook(filepath)
    ws = wb.get_sheet_by_name(sheet_name)

    # CREATE FILTER
    # ws.auto_filter.add_filter_column(0, list_columns, blank=False)
    ws.auto_filter.ref = ws.dimensions

    # "UNCLEAR"
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf_red = DifferentialStyle(fill=red_fill)
    r_red = Rule(type="expression", dxf=dxf_red, stopIfTrue=True)
    r_red.formula = ['$B1="U"']

    # "CLOSED"
    green_fill = PatternFill(bgColor="D0FFC7")
    dxf_green = DifferentialStyle(fill=green_fill)
    r_green = Rule(type="expression", dxf=dxf_green, stopIfTrue=True)
    r_green.formula = ['$B1="C"']

    # "OPENED"
    orange_fill = PatternFill(bgColor="FFA54C")
    dxf_orange = DifferentialStyle(fill=orange_fill)
    r_orange = Rule(type="expression", dxf=dxf_orange, stopIfTrue=True)
    r_orange.formula = ['$B1="O"']

    # APPLY
    ws.conditional_formatting.add("A1:P10000", r_red)
    ws.conditional_formatting.add("A1:P10000", r_green)
    ws.conditional_formatting.add("A1:P10000", r_orange)

    # WRAP HEADER
    for header in list(string.ascii_uppercase[:dt.shape[0]]):
        ws['{}1'.format(header)].alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(filepath)


def xlsx_create_overview(filepath, sheet_name="overview"):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.create_sheet(sheet_name)

        # ADD VALUES TO CELLS
        ws['A1'] = "Profit brutto:"
        ws['A2'] = "Fees:"
        ws['A3'] = "Profit netto:"
        ws['A5'] = "Date from:"
        ws['A6'] = "Date to:"
        ws['A7'] = "Date from (unix):"
        ws['A8'] = "Date to (unix):"

        ws['B1'] = '=SUMIFS(full_data!O1:O100000, full_data!P1:P100000, "CLOSED", full_data!D1:D100000, ">="&B7, full_data!D1:D100000, "<"&B8, full_data!B1:B100000, "C")'
        ws['B2'] = '=SUMIFS(full_data!W1:W100000, full_data!D1:D100000, ">="&B7, full_data!D1:D100000, "<"&B8)'
        ws['B3'] = '=B1-B2'
        ws['B5'] = "01/01/2021"
        ws['B6'] = "01/01/2022"
        ws['B7'] = "=(B5-DATE(1970,1,1))*86400"
        ws['B8'] = "=(B6-DATE(1970,1,1))*86400"

        ws["C1"] = "USD"
        ws["C2"] = "USD"
        ws["C3"] = "USD"

        # SET WIDTH OF COLUMNS
        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 11

        # SET FORMAT OF CELLS/COLUMNS
        for row in range(1, 4):
            ws["B{}".format(row)].number_format = '#,##0.00'
            
            ws["A{}".format(row)].font = Font(bold=True)
            ws["B{}".format(row)].font = Font(bold=True)
            ws["C{}".format(row)].font = Font(bold=True)
        
        # HIDE SHEET WITH ALL DATA
        ws2 = wb.get_sheet_by_name("full_merge")
        ws2.sheet_state = "hidden"

        wb.save(filepath)


def xlsx_export(filepath, dt, network, sheet_name="full_data"):
    # IF FILE EXISTS, DELETE IT
    if os.path.exists(filepath):
        os.remove(filepath)
    
    with pd.ExcelWriter(filepath, engine="xlsxwriter") as writer:
        dt.to_excel(writer, sheet_name=sheet_name, index=False, freeze_panes=(1,1))
        # Get the xlsxwriter objects from the dataframe writer object:
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        # Hide columns:
        dict_columns_to_hide = adjust_dict(_dict_columns_to_hide, network)
        for col_name in dict_columns_to_hide.values():
            worksheet.set_column("{}:{}".format(col_name, col_name), None, None, {"hidden": 1})
        # Modify columns:
        for col_name, width, bold in dict_columns_width.values():
            cell_format = workbook.add_format({"bold": bold})
            worksheet.set_column("{}:{}".format(col_name, col_name), width, cell_format)


def xlsx_export_data_to_sheet(filepath, dt, sheet_name="full_merge"):
    if os.path.exists(filepath):
        with pd.ExcelWriter(filepath, engine="openpyxl", mode="a") as writer:
            dt.to_excel(writer, sheet_name=sheet_name, index=False, freeze_panes=(1,1))


def xlsx_open(filepath):
    os.system("start EXCEL.EXE {}".format(filepath))